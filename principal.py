import os
import errno
import openpyxl
import numpy
import random
import validar
from docxtpl import InlineImage
from docxtpl import DocxTemplate

class Princpal:
    def __init__(self):
        try:
            os.mkdir('C://Users/'+os.getlogin()+'/Desktop/Documentos_IINOVA')
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise
        self.empresas=[]
        self.trabajo=[]
        self.numEm=1
        self.numT=1
        self.numCont=0
        self.conEmp=[]
        self.contrato=[]
        self.f=''
        self.fecha=''
        self.obra=[]
        self.convo_esta=''
        self.num_lic=[]
        self.contG=1
        self.numL=1
        self.cartas=0
        print('\t==================INGENIERIA INOVA==================')
        print('Software para la elaboracion de documentos tecnicos y economicos')
        print('Creado por: Luis Mario Rosales Medina')
        print()
        print('CONSIDERACIONES')
        print('1.-La carpeta informacion y el ejecutable deben estar siempre juntos para que el programa funcione')
        print('2-.Si se desea añadir una nueva empresa, se debe añadir tanto al excel (INOVA.xlsx) la informacion, como tambien los pies y cabecera (Imagenes) en la carpeta "informacion"')
        print('3.-El nombre de la empresa a utilizar para añadir pie de pagina y cabecera es el que usted ponga en la columna "C" de la hoja "Empresas" del Excel INOVA')
        print('4-.Si desea actualizar informacion de alguna empresa o trabajo, unicamente modifique el campo requerido  en el excel')
        print('5.-Evite modificar en los templates de word (formatos) las partes o expresiones que estan contenidas entre doble corchete {{algo}} pues es aqui donde el programa vacia la informacion')
        print('6.-Evite dejar campos vacios si modifica informacion sobre las empresas. (Si no posee la informacion ponga cualquier cosa)')
        input()

    def cam_Hoja(self,clave,rango,ruta):
        os.system("cls")
        self.excel=openpyxl.load_workbook(ruta)
        self.excel.active=clave
        self.hoja=self.excel.active
        self.datos=numpy.transpose(self.hoja[rango])
        for i in self.datos:
            contador=0
            for j in i:
                print(str(j.value)+'\t',end=' ')
                contador+=1
                if contador==3:
                    break
            print(' ')
    
    def cam_Hoja2(self):
        self.excel=openpyxl.load_workbook('Informacion\EXPERIENCIA.xlsx')
        j=0
        while(j<len(self.empresas)):
            os.system("cls")
            self.excel.active=self.empresas[j][0].value-1
            self.hoja=self.excel.active
            self.datos=numpy.transpose(self.hoja['A:G'])
            print("Contratos de "+self.empresas[j][1].value)
            for i in self.datos:
                print(str(i[0].value)+'\t'+str(i[1].value)+'\t'+str(i[2].value),end=' ')
                print(' ')
                print(' ')
                print("No. de contratos seleccionados:"+str(len(self.contrato)))
            self.numCont=validar.Validar.ValEnt('Que contrato deseas (0 si no quiere ninguno):',len(self.datos))
            if (self.numCont>0):
                self.contrato.append(self.datos[self.numCont-1])
                self.conEmp.append(self.empresas[j][1].value)
            else:
                j+=1



    def get_Trab(self):
        self.cam_Hoja(0,'A:J','Informacion\IINOVA.xlsx')
        self.numT=validar.Validar.ValEnt('Que trabajo deseas (coloca 0 si ya no quieres mas):',len(self.datos))
        if self.numT==0:
            exit()
        self.trabajo.append(self.datos[self.numT-1])
        self.info1()
        self.get_empresas(0)
        
    def get_empresas(self,aux):
        respuesta=0
        while self.numEm!=0:
            os.system("cls")
            self.cam_Hoja(1,'A:P','Informacion\IINOVA.xlsx')
            if aux==0:
                self.numEm=validar.Validar.ValEnt('De que empresas quieres (coloca 0 si ya no quieres mas):',len(self.datos))
                if self.numEm!=0:
                    self.empresas.append(self.datos[self.numEm-1])
                    if self.numT==1 or self.numT==2 or self.numT==5 or self.numT==6 or self.numT==9 or self.numT==10 or self.numT==13 or self.numT==14 or self.numT==17 or self.numT==18 or self.numT==22 or self.numT==23:
                        respuesta=validar.Validar.ValEle('¿Va en asociacion? (1=si, 0=no):')
                    if respuesta==1:
                        self.get_empresas(1)
                        self.numEm=-1
                    self.selector()
                    self.empresas.clear()
                    self.contG+=1
            elif aux==1:
                self.numEm=validar.Validar.ValNum('Cual es la empresa asociada (coloca 0 si ya no quieres mas):')
                if self.numEm!=0:
                    self.empresas.append(self.datos[self.numEm-1])
                
                
        os.system("cls")

    def selector(self):
        if self.numT>0 & self.numT<=24:
            self.licitaciones()


    def info1(self):#Informacion general del documento 
        self.f=validar.Validar.ValGen('Dame la fecha de entrega (DD de MMMM DEL YYYY):')
        if self.numT<=8 or self.numT>=17 or self.numT==100:
            self.fecha='Aguascalientes, Ags.,'+self.f
        elif self.numT>=9 and self.numT<=12:
            self.fecha='Rincon de Romos, Ags.,'+self.f
        elif self.numT>=13 and self.numT<=16:
            self.fecha='Jesus Maria, Ags.,'+self.f
        if self.numT!=21:
            self.convo_esta=validar.Validar.ValGen('Dame la clave de la conocatoria:')
        self.constantes={'fecha':self.fecha,'destinatario':self.trabajo[0][3].value,'cargo_des':self.trabajo[0][4].value,
                        'convo_esta':self.convo_esta,'fecha1':int(self.fecha[-4:])-3,'fecha2':int(self.fecha[-4:])-2,'fecha3':int(self.fecha[-4:])-1}

        if self.numT==1 or self.numT==2 or self.numT==5 or self.numT==6 or self.numT==9 or self.numT==10 or self.numT==13 or self.numT==14 or self.numT==17 or self.numT==18 or self.numT==22 or self.numT==23 or self.numT==24:
            self.constantes['tipo']='LA CONVOCATORIA'
        else:
            self.constantes['tipo']='LA INVITACIÓN'

        if self.numT<=4:
            self.constantes['ente']='LA SECRETARIA'
        else:
            self.constantes['ente']='EL INSTITUTO'

        if self.numT==1 or self.numT==5 or self.numT==9 or self.numT==13 or self.numT==17 or self.numT==22:
            self.numL=validar.Validar.ValNum('A cuantas obras se licita las empresas:')
            for i in range(self.numL):
                self.num_lic.append(validar.Validar.ValGen('Dame el numero de licitacion o bien el paquete de la obra:'))
                self.constantes['num_lic'+str(i+1)]=self.num_lic[i]
                self.obra.append(validar.Validar.ValGen('Dame el nombre del la obra:'))
                self.constantes['nom_obra'+str(i+1)]=self.obra[i]
            self.constantes['monto']=validar.Validar.ValFlo('Dame el monto de la obra:')
        else:
            self.num_lic.append(validar.Validar.ValGen('Dame el numero de licitacion:'))
            self.constantes['num_lic1']=self.num_lic[0]
            self.obra.append(validar.Validar.ValGen('Dame el nombre del la obra:'))
            self.constantes['nom_obra1']=self.obra[0]
            if self.numT!=24:
                self.constantes['subcontrato']=validar.Validar.ValGen('Dame lo que se subcontrata:')
                self.constantes['tiie1']=validar.Validar.ValFlo('Dame la TIIE:')
                self.constantes['tiie2']=round(15-float(self.constantes['tiie1']),4)
                if self.numT!=21:
                    self.constantes['cargoT']=0
                    for i in range(5):
                        self.constantes['cargo'+str(i+1)]=float(self.trabajo[0][i+5].value)
                        self.constantes['cargoT']=round(self.constantes['cargoT']+self.constantes['cargo'+str(i+1)],2)

        self.cartas=validar.Validar.ValEle('¿Necesitas cartas de concreto, aslfato etc? (1=Si 0=No):')

        if self.numT!=21:
            self.nueva_ruta='C://Users/'+os.getlogin()+'/Desktop/Documentos_IINOVA/'+self.convo_esta+'_'+self.num_lic[0]+'/'
        else:
            self.constantes.pop('convo_esta')
            self.constantes.pop('cargo_des')
            self.nueva_ruta='C://Users/'+os.getlogin()+'/Desktop/Documentos_IINOVA/'+'INEGI_'+self.num_lic[0]+'/'
    
    def info2(self):#Informacion esepcializada de la empresa
        for j in range(1,len(self.empresas)):
                self.constantes['empresa'+str(j+1)]=self.empresas[j][1].value
        for i in range(len(self.empresas)):
            if self.numT==1 or self.numT==5 or self.numT==9 or self.numT==13 or self.numT==17 or self.numT==22:
                self.doc=DocxTemplate('Informacion\Words\Molde1_V'+str(self.contG)+'.docx')
            elif self.numT==24:
                self.doc=DocxTemplate('Informacion\Words\Molde4_V'+str(self.contG)+'.docx')
            else:
                self.doc=DocxTemplate('Informacion\Words\Molde2_V'+str(self.contG)+'.docx')
                if i==0:
                    self.constantes['utilidad']=validar.Validar.ValFlo("Dame la utilidad de "+self.empresas[i][2].value+":")
            self.constantes['empresa1']=self.empresas[i][1].value
            self.constantes['rep_leg']=self.empresas[i][3].value
            self.constantes['calle']=self.empresas[i][5].value
            self.constantes['colonia']=self.empresas[i][6].value
            self.constantes['postal']=self.empresas[i][7].value
            self.constantes['municipio']=self.empresas[i][8].value
            self.constantes['estado']=self.empresas[i][9].value
            self.constantes['telefono']=self.empresas[i][10].value
            self.constantes['correo_emp']=self.empresas[i][11].value
            self.constantes['residente']=self.empresas[i][4].value
            self.constantes['rfc_E']=self.empresas[i][14].value###################
            self.constantes['rfc_R']=self.empresas[i][15].value###################
            if os.path.isfile('Informacion/Imagenes/'+self.empresas[i][2].value+'1.jpg'):
                self.constantes['cabecera']=InlineImage(self.doc,'Informacion/Imagenes/'+self.empresas[i][2].value+'1.jpg')
            if os.path.isfile('Informacion/Imagenes/'+self.empresas[i][2].value+'2.jpg'):
                self.constantes['pie']=InlineImage(self.doc,'Informacion/Imagenes/'+self.empresas[i][2].value+'2.jpg')
            if self.numT!=21:
                self.doc.render(self.constantes)
                self.doc.save(self.nueva_ruta+self.empresas[i][2].value+'_'+self.convo_esta+'_'+self.num_lic[0]+'_'+self.trabajo[0][1].value+'.docx')
            else:
                #print(self.constantes.items())
                for k in range(3):
                    self.doc=DocxTemplate('Informacion\Words\Molde3_P'+str(k+1)+'.docx')
                    if os.path.isfile('Informacion/Imagenes/'+self.empresas[i][2].value+'1.jpg'):
                        self.constantes['cabecera']=InlineImage(self.doc,'Informacion/Imagenes/'+self.empresas[i][2].value+'1.jpg')
                    if os.path.isfile('Informacion/Imagenes/'+self.empresas[i][2].value+'2.jpg'):
                        self.constantes['pie']=InlineImage(self.doc,'Informacion/Imagenes/'+self.empresas[i][2].value+'2.jpg')
                    #self.doc=DocxTemplate('Informacion\Words\Molde3_P3.docx')
                    self.doc.render(self.constantes)
                    self.doc.save(self.nueva_ruta+self.empresas[i][2].value+'_INEGI_'+self.num_lic[0]+'_'+str(k+1)+'.docx')
    
    def info3(self): #Información excel
        if self.numT==1:
            for j in range(self.numL): 
                self.cam_Hoja2()###################################### econtrato puede variar si son varias licitaciones
                self.excel=openpyxl.load_workbook('Informacion\Excels\Molde1.xlsx')
                self.excel.active=0
                self.hoja=self.excel.active
                for i in range(len(self.empresas)):
                    self.hoja['C'+str(8+i)]=self.empresas[i][1].value
                    self.hoja['F'+str(8+i)]=self.empresas[i][13].value
                    self.hoja['G'+str(8+i)]=self.empresas[i][12].value

                for i in range(len(self.contrato)):
                    self.hoja['B'+str(22+i)]=self.contrato[i][6].value
                    self.hoja['C'+str(22+i)]=self.contrato[i][3].value
                    self.hoja['E'+str(22+i)]=self.contrato[i][1].value
                    self.hoja['H'+str(22+i)]=self.contrato[i][2].value
                    self.hoja['I'+str(22+i)]=str(self.contrato[i][4].value)+' al '+str(self.contrato[i][5].value)
                self.contrato.clear()
                self.conEmp.clear()

                self.hoja['C17']=self.num_lic[j]
                self.hoja['D17']=self.obra[j]
                self.excel.active=4
                self.hoja=self.excel.active
                self.hoja['E16']=self.f
                self.hoja['D19']=self.empresas[0][3].value
                self.hoja['C28']=self.convo_esta
                self.excel.save(self.nueva_ruta+self.empresas[0][2].value+'_'+self.convo_esta+'_'+self.num_lic[j]+'_'+self.trabajo[0][1].value+'.xlsx')
                self.excel.close()
        elif self.numT==2 or self.numT==3 or self.numT==4 or self.numT==9 or self.numT==10 or self.numT==11 or self.numT==12 or self.numT==22 or self.numT==23:
            for j in range(self.numL):
                self.cam_Hoja2()######
                self.excel=openpyxl.load_workbook('Informacion\Excels\Molde2.xlsx')
                self.excel.active=0
                self.hoja=self.excel.active
                for i in range(len(self.empresas)):
                    self.hoja['C'+str(8+i)]=self.empresas[i][1].value
                    self.hoja['F'+str(8+i)]=self.empresas[i][13].value
                self.hoja['C17']=self.num_lic[j]
                self.hoja['D17']=self.obra[j]
                self.hoja['C15']=self.f
                self.hoja['D19']=self.empresas[0][3].value

                self.excel.active=1
                self.hoja=self.excel.active
                for i in range(len(self.contrato)):
                    self.hoja['B'+str(13+i)]=self.contrato[i][1].value
                    self.hoja['C'+str(13+i)]=self.contrato[i][2].value
                    self.hoja['D'+str(13+i)]=self.contrato[i][3].value
                    self.hoja['E'+str(13+i)]=self.conEmp[i]
                    self.hoja['F'+str(13+i)]=str(self.contrato[i][4].value)+' al '+str(self.contrato[i][5].value)
                    self.hoja['G'+str(13+i)]=self.contrato[i][1].value
                    self.hoja['H'+str(13+i)]='RESIDENTE'
                    self.hoja['I'+str(13+i)]=self.contrato[i][6].value
                self.excel.active=2
                self.hoja=self.excel.active
                for i in range(len(self.contrato)):
                    self.hoja['B'+str(13+i)]=self.contrato[i][6].value
                    self.hoja['C'+str(13+i)]=self.contrato[i][3].value
                    self.hoja['D'+str(13+i)]=self.contrato[i][1].value
                    self.hoja['E'+str(13+i)]=self.contrato[i][2].value
                    self.hoja['F'+str(13+i)]=str(self.contrato[i][4].value)
                    self.hoja['G'+str(13+i)]=str(self.contrato[i][5].value)
                self.contrato.clear()
                self.conEmp.clear()

                self.excel.save(self.nueva_ruta+self.empresas[0][2].value+'_'+self.convo_esta+'_'+self.num_lic[j]+'_'+self.trabajo[0][1].value+'.xlsx')
                self.excel.close()
        elif self.numT>=5 and self.numT<=8:
            for j in range(self.numL):
                self.cam_Hoja2()#######
                self.excel=openpyxl.load_workbook('Informacion\Excels\Molde3.xlsx')
                self.excel.active=0
                self.hoja=self.excel.active
                for i in range(len(self.empresas)):
                    self.hoja['C'+str(8+i)]=self.empresas[i][1].value
                    self.hoja['F'+str(8+i)]=self.empresas[i][13].value
                    self.hoja['G'+str(8+i)]=self.empresas[i][12].value
                self.excel.active=1
                self.hoja=self.excel.active
                self.hoja['E10']=self.f
                self.hoja['E11']=self.empresas[0][1].value
                self.hoja['E12']=self.obra[j]
                self.hoja['E13']=self.num_lic[j]
                self.hoja['E14']=self.empresas[0][12].value

                self.excel.active=5
                self.hoja=self.excel.active
                for i in range(len(self.contrato)):
                    self.hoja['C'+str(17+i)]=self.contrato[i][6].value
                    self.hoja['D'+str(17+i)]=self.contrato[i][3].value
                    self.hoja['G'+str(17+i)]=self.contrato[i][1].value
                    self.hoja['I'+str(17+i)]=self.contrato[i][2].value
                    self.hoja['J'+str(17+i)]=str(self.contrato[i][4].value)
                    self.hoja['K'+str(17+i)]=str(self.contrato[i][5].value)
                self.excel.active=6
                self.hoja=self.excel.active
                for i in range(len(self.contrato)):
                    self.hoja['C'+str(15+i)]=self.contrato[i][1].value
                    self.hoja['D'+str(15+i)]=self.contrato[i][2].value
                    self.hoja['E'+str(15+i)]=self.contrato[i][3].value
                    self.hoja['F'+str(15+i)]=self.conEmp[i]
                    self.hoja['G'+str(15+i)]=str(self.contrato[i][4].value)+' al '+str(self.contrato[i][5].value)
                    self.hoja['H'+str(15+i)]=self.contrato[i][1].value
                    self.hoja['I'+str(15+i)]='RESIDENTE'
                    self.hoja['J'+str(15+i)]=self.contrato[i][6].value
                self.contrato.clear()
                self.conEmp.clear()

                self.excel.save(self.nueva_ruta+self.empresas[0][2].value+'_'+self.convo_esta+'_'+self.num_lic[j]+'_'+self.trabajo[0][1].value+'.xlsx')
                self.excel.close()
        elif self.numT>=13 and self.numT<=16:
            for j in range(self.numL):
                self.cam_Hoja2()#######
                self.excel=openpyxl.load_workbook('Informacion\Excels\Molde4.xlsx')
                self.excel.active=0
                self.hoja=self.excel.active
                for i in range(len(self.empresas)):
                    self.hoja['C'+str(2+i)]=self.empresas[i][1].value
                    self.hoja['G'+str(2+i)]=self.empresas[i][13].value
                self.hoja['C9']=self.f
                self.hoja['C7']=self.obra[j]
                self.hoja['C8']=self.num_lic[j]

                self.excel.active=3
                self.hoja=self.excel.active
                for i in range(len(self.contrato)):
                    self.hoja['A'+str(16+i)]=self.contrato[i][6].value
                    self.hoja['B'+str(16+i)]=self.contrato[i][3].value
                    self.hoja['D'+str(16+i)]=self.contrato[i][1].value
                    self.hoja['F'+str(16+i)]=self.contrato[i][2].value
                    self.hoja['G'+str(16+i)]=str(self.contrato[i][4].value)+' al '+str(self.contrato[i][5].value)
                self.excel.active=4
                self.hoja=self.excel.active
                for i in range(len(self.contrato)):
                    self.hoja['B'+str(16+i)]=self.contrato[i][1].value
                    self.hoja['C'+str(16+i)]=self.contrato[i][2].value
                    self.hoja['D'+str(16+i)]=self.contrato[i][3].value
                    self.hoja['E'+str(16+i)]=self.conEmp[i]
                    self.hoja['F'+str(16+i)]=str(self.contrato[i][4].value)+' al '+str(self.contrato[i][5].value)
                    self.hoja['G'+str(16+i)]=self.contrato[i][1].value
                    self.hoja['H'+str(16+i)]='RESIDENTE'
                    self.hoja['I'+str(16+i)]=self.contrato[i][6].value
                self.contrato.clear()
                self.conEmp.clear()

                self.excel.save(self.nueva_ruta+self.empresas[0][2].value+'_'+self.convo_esta+'_'+self.num_lic[j]+'_'+self.trabajo[0][1].value+'.xlsx')
                self.excel.close()
        elif self.numT>=17 and self.numT<=20:
            for j in range(self.numL):
                self.cam_Hoja2()########
                self.excel=openpyxl.load_workbook('Informacion\Excels\Molde5.xlsx')
                self.excel.active=0
                self.hoja=self.excel.active
                for i in range(len(self.empresas)):
                    self.hoja['B'+str(11+i)]=self.empresas[i][1].value
                    self.hoja['G'+str(11+i)]=self.empresas[i][13].value
                #self.hoja['b10']=self.f
                self.hoja['B16']=self.obra[j]
                self.hoja['B17']=self.num_lic[j]

                self.excel.active=1
                self.hoja=self.excel.active
                for i in range(len(self.contrato)):
                    self.hoja['A'+str(14+i)]=self.contrato[i][1].value
                    self.hoja['B'+str(14+i)]=self.contrato[i][2].value
                    self.hoja['C'+str(14+i)]=self.contrato[i][3].value
                    self.hoja['D'+str(14+i)]=self.conEmp[i]
                    self.hoja['E'+str(14+i)]=str(self.contrato[i][4].value)+' al '+str(self.contrato[i][5].value)
                    self.hoja['F'+str(14+i)]=self.contrato[i][1].value
                    self.hoja['G'+str(14+i)]='RESIDENTE'
                    self.hoja['H'+str(14+i)]=self.contrato[i][6].value
                self.excel.active=2
                self.hoja=self.excel.active
                for i in range(len(self.contrato)):
                    self.hoja['A'+str(14+i)]=self.contrato[i][6].value
                    self.hoja['B'+str(14+i)]=self.contrato[i][3].value
                    self.hoja['D'+str(14+i)]=self.contrato[i][1].value
                    self.hoja['G'+str(14+i)]=self.contrato[i][2].value
                    self.hoja['H'+str(14+i)]=str(self.contrato[i][4].value)+' al '+str(self.contrato[i][5].value)
                self.contrato.clear()

                self.excel.save(self.nueva_ruta+self.empresas[0][2].value+'_'+self.convo_esta+'_'+self.num_lic[j]+'_'+self.trabajo[0][1].value+'.xlsx')
                self.excel.close()

    
    def info4(self): #Maquinaria, Capacidad tecnica y Contrato de asociacion
        if self.numT!=1 and self.numT!=5 and self.numT!=9 and self.numT!=13 and self.numT!=17 and self.numT!=22 and self.numT!=24:
            if(len(self.empresas)>1):
                self.doc=DocxTemplate('Informacion\Words\Contrato_Asociacion.docx')
                for i in range(len(self.empresas)):
                    self.constantes['empresa'+str(i+1)]=self.empresas[i][1].value
                    self.constantes['rep_leg'+str(i+1)]=self.empresas[i][3].value
                    if (i>0):
                        self.constantes['calle'+str(i+1)]=self.empresas[i][5].value
                        self.constantes['colonia'+str(i+1)]=self.empresas[i][6].value
                        self.constantes['postal'+str(i+1)]=self.empresas[i][7].value
                        self.constantes['municipio'+str(i+1)]=self.empresas[i][8].value
                self.doc.render(self.constantes)
                self.doc.save(self.nueva_ruta+self.empresas[0][2].value+'_'+self.convo_esta+'_'+self.num_lic[0]+'_Contrato de asociacion.docx')
            self.doc=DocxTemplate('Informacion\Words\Maquinaria'+str(random.randint(1,3))+'.docx')
            self.doc.render(self.constantes)
            self.doc.save(self.nueva_ruta+self.empresas[0][2].value+'_'+self.convo_esta+'_'+self.num_lic[0]+'_Maquinaria.docx')
        if (self.cartas==1):
            self.doc=DocxTemplate('Informacion\Words\Concreto'+str(random.randint(1,3))+'.docx')
            self.doc.render(self.constantes)
            self.doc.save(self.nueva_ruta+self.empresas[0][2].value+'_'+self.convo_esta+'_'+self.num_lic[0]+'_Concreto.docx')
            self.doc=DocxTemplate('Informacion\Words\Asfalto'+str(random.randint(1,2))+'.docx')
            self.doc.render(self.constantes)
            self.doc.save(self.nueva_ruta+self.empresas[0][2].value+'_'+self.convo_esta+'_'+self.num_lic[0]+'_Asfalto.docx')


    def licitaciones(self):
        try:
            os.mkdir(self.nueva_ruta)
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise

        self.info2()
        self.info3()
        self.info4()
        #self.info5()
        if self.contG==5:
            self.contG=0




